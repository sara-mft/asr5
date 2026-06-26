"""
reporters.py — Output Formatters
========================================
"""

from __future__ import annotations

import abc
import csv
import json
import logging
import random
from collections import defaultdict
from pathlib import Path
from typing import Sequence

from models import TranscriptionResult

log = logging.getLogger(__name__)



# ──────────────────────────────────────────────────────────────────────────────
# Base class
# ──────────────────────────────────────────────────────────────────────────────


class BaseReporter(abc.ABC):
    @abc.abstractmethod
    def write(self, results: Sequence[TranscriptionResult]) -> None:
        """Persist or display the benchmark results."""


# ──────────────────────────────────────────────────────────────────────────────
# JSON reporter
# ──────────────────────────────────────────────────────────────────────────────


class JSONReporter(BaseReporter):
    """Writes structured JSON with per-engine aggregate statistics."""

    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path

    def write(self, results: Sequence[TranscriptionResult]) -> None:
        serialised = []
        for r in results:
            row = r.to_dict()
            if r.has_diarization:
                row["diarization_segments"] = [s.to_dict() for s in r.diarization_segments]
            serialised.append(row)

        output = {
            "summary": _build_summary(results),
            "results": serialised,
        }
        self.output_path.write_text(
            json.dumps(output, indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
        log.info("JSON results written to: %s", self.output_path)


# ──────────────────────────────────────────────────────────────────────────────
# CSV reporter
# ──────────────────────────────────────────────────────────────────────────────


class CSVReporter(BaseReporter):
    """Flat CSV — one row per (engine × file) cell."""

    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path

    def write(self, results: Sequence[TranscriptionResult]) -> None:
        if not results:
            log.warning("No results to write to CSV.")
            return
        rows = [r.to_dict() for r in results]
        fieldnames = list(rows[0].keys())
        with open(self.output_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        log.info("CSV results written to: %s", self.output_path)


# ──────────────────────────────────────────────────────────────────────────────
# Excel reporter
# ──────────────────────────────────────────────────────────────────────────────


class ExcelReporter(BaseReporter):
    """Polished Excel workbook: Summary Dashboard + Detailed Results tabs."""

    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path

    def write(self, results: Sequence[TranscriptionResult]) -> None:
        try:
            import pandas as pd
            # Ensure Workbook is imported from openpyxl
            from openpyxl import Workbook 
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            log.error(
                "pandas or openpyxl not installed. Run: pip install pandas openpyxl"
            )
            return

        if not results:
            log.warning("No results to write to Excel.")
            return

        summary = _build_summary(results)
        
        summary_rows = []
        for eid, stats in summary.items():
            row = stats.copy()

            summary_rows.append({"Engine": eid, **row})

        import pandas as pd
        df_summary = pd.DataFrame(summary_rows)
        df_detailed = pd.DataFrame([r.to_dict() for r in results])

        # 2. Now initialize the Workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary Dashboard"
        ws_detailed = wb.create_sheet(title="Detailed Results")

        # Percentage columns (format as 0.00%)
        ## start modif ##
        PCT_KEYS = {
            "wer", "cer", "mer", "wip", "wil", "tner", "har",
            "norm_wer", "rate", "onset_wer", "jer"
        } ## diar modif ##
        ## end modif ##


        # Currency columns
        USD_KEYS = {"usd", "cost"}

        def _is_pct(col_name: str) -> bool:
            return any(k in col_name.lower() for k in PCT_KEYS)

        def _is_usd(col_name: str) -> bool:
            return any(k in col_name.lower() for k in USD_KEYS)

        def format_sheet(ws, df) -> None:
            header_fill  = PatternFill("solid", fgColor="2F4F4F")
            alt_fill     = PatternFill("solid", fgColor="F9F9F9")
            header_font  = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D3D3D3")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(row)
                for c_idx, cell in enumerate(ws[r_idx], 1):
                    cell.border = border
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center")
                        if r_idx % 2 == 0:
                            cell.fill = alt_fill
                        col_name = str(df.columns[c_idx - 1])
                        if isinstance(cell.value, float):
                            if _is_pct(col_name):
                                cell.number_format = "0.00%"
                            elif _is_usd(col_name):
                                cell.number_format = '$#,##0.0000'
                            elif any(k in col_name.lower() for k in ("rtf", "latency", "f1", "pper")):
                                cell.number_format = "0.0000"

            # Auto-fit columns
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=10,
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            ws.auto_filter.ref = ws.dimensions

        format_sheet(ws_summary, df_summary)
        ws_summary.freeze_panes = "A2"
        format_sheet(ws_detailed, df_detailed)
        ws_detailed.freeze_panes = "C2"

        wb.save(self.output_path)
        log.info("Excel results written to: %s", self.output_path)


# ──────────────────────────────────────────────────────────────────────────────
# Console reporter
# ──────────────────────────────────────────────────────────────────────────────


class ConsoleReporter(BaseReporter):
    """
    Formatted summary table printed to stdout.

    """

    def write(self, results: Sequence[TranscriptionResult]) -> None:
        summary = _build_summary(results)
        if not summary:
            print("No results to display.")
            return

        ## start modif ##
        col_w   = [22, 5, 7, 6, 5, 7, 8, 7, 7, 7, 7, 7, 9]
        headers = ["Engine", "Files", "Lat(s)", "TTFW","RTF",  "WER", "NormWER",  "OnsetWER" , "MER", "WIP", "HAR", "TNER", "Cost"]
        ## end modif ##

        ## diar modif ##
        col_w   = [22, 5, 7, 6, 5, 7, 8, 7, 7, 7, 7, 7, 7, 7, 9]
        headers = ["Engine", "Files", "Lat(s)", "TTFW","RTF",  "WER", "NormWER",  "OnsetWER" , "MER", "WIP", "HAR", "TNER", "DER", "JER","Cost"]
        ## diar modif ##

        sep = "─" * (sum(col_w) + len(col_w) * 3 + 1)

        print()
        print("  Azure STT Benchmark v2 — Summary")
        print(f"  {sep}")
        print("  " + " │ ".join(h.ljust(w) for h, w in zip(headers, col_w)))
        print(f"  {sep}")

        for engine_id, stats in summary.items():
            def fmt_pct(val):
                return f"{val:.1%}" if val is not None else "N/A"

            row_vals = [
                engine_id[:col_w[0]],
                str(stats["num_files"]),
                f"{stats['avg_latency_sec']:.1f}" if stats["avg_latency_sec"] is not None else "N/A",
                ## start modif ##
                f"{stats['avg_ttfw_sec']:.2f}"    if stats.get("avg_ttfw_sec") is not None else "N/A",
                ## end modif ##
                f"{stats['avg_rtf']:.2f}"          if stats["avg_rtf"] is not None else "N/A",
                fmt_pct(stats["avg_wer"]),
                fmt_pct(stats["avg_norm_wer"]),
                ## start modif ##
                fmt_pct(stats["avg_onset_wer"]),
                ## end modif ##
                fmt_pct(stats["avg_mer"]),
                fmt_pct(stats["avg_wip"]),
                fmt_pct(stats["avg_har"]),
                fmt_pct(stats["avg_tner"]),
                fmt_pct(stats["avg_der"]),  ## diar modif ##
                fmt_pct(stats["avg_jer"]),  ## diar modif ##
                f"${stats['total_cost_usd']:.4f}",
            ]
            print("  " + " │ ".join(v.ljust(w) for v, w in zip(row_vals, col_w)))

        print(f"  {sep}")


        # Diarization transcripts
        diar_results = [r for r in results if r.has_diarization]
        if diar_results:
            print()
            print("  ── Diarization Transcripts " + "─" * 60)
            for r in diar_results:
                print(f"\n  [{r.engine_id}]  {r.audio_file}  "
                      f"({r.metrics.num_speakers} speaker(s))\n")
                for line in r.diarization_transcript.splitlines():
                    print(f"    {line}")
        print()


# ──────────────────────────────────────────────────────────────────────────────
# Shared aggregation 
# ──────────────────────────────────────────────────────────────────────────────



def _build_summary(
    results: Sequence[TranscriptionResult],
) -> dict:
    buckets: dict[str, list[TranscriptionResult]] = defaultdict(list)
    for r in results:
        buckets[r.engine_id].append(r)

    summary = {}
    for engine_id, rows in buckets.items():
        successful = [r for r in rows if r.success]
        n = len(successful)

        def get_avg(attr: str) -> float | None:
            vals = [
                v for r in successful
                if (v := getattr(r.metrics, attr)) is not None
            ]
            return _safe_mean(vals)

        wer_vals = [
            r.metrics.word_error_rate
            for r in successful
            if r.metrics.word_error_rate is not None
        ]

        entry: dict = {
            "num_files": len(rows),
            "num_successful": n,
            "num_errors": len(rows) - n,

            # Performance & cost
            "avg_latency_sec": round(get_avg("latency_sec") or 0.0, 4),
            ## start modif ##
            "avg_ttfw_sec":    get_avg("time_to_first_token_sec"),  
            ## end modif ##

            "avg_rtf":         round(get_avg("real_time_factor") or 0.0, 4),
            "total_cost_usd":  round(
                sum(r.metrics.estimated_cost_usd for r in successful), 6
            ),
            "avg_confidence":  get_avg("confidence"),
            "avg_speakers":    (
                round(s) if (s := get_avg("num_speakers")) is not None else None
            ),

            # Standard accuracy
            "avg_wer":      get_avg("word_error_rate"),
            ## start modif ##
            "avg_onset_wer": get_avg("onset_wer"),  
            ## end modif ##
            "avg_onset_wer": get_avg("onset_wer"),
            "avg_cer":      get_avg("character_error_rate"),
            "avg_mer":      get_avg("match_error_rate"),
            "avg_wip":      get_avg("word_info_preserved"),
            "avg_wil":      get_avg("word_info_lost"),
            "avg_norm_wer": get_avg("norm_word_error_rate"),

            # Domain / semantic
            "avg_tner":     get_avg("term_error_rate"),
            "avg_har":      get_avg("hallucination_rate"),

            # Punctuation & readability
            "avg_punc_f1":  get_avg("punctuation_f1_macro"),
            "avg_pper":     get_avg("punctuation_error_rate"),

            ## diar modif ##
            # Diarization
            "avg_der":      get_avg("diarization_error_rate"),
            "avg_jer":      get_avg("jaccard_error_rate"),
            ## diar modif ##

        }
        summary[engine_id] = entry

    return summary


def _safe_mean(values: list[float]) -> float | None:
    return sum(values) / len(values) if values else None
